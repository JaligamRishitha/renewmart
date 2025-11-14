from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from fastapi.responses import Response, StreamingResponse
from fastapi.security import HTTPAuthorizationCredentials
from sqlalchemy.orm import Session
from sqlalchemy import text
from typing import List, Optional, Dict, Any
from uuid import UUID
from decimal import Decimal
from datetime import datetime
import json
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from database import get_db
from auth import get_current_user, require_admin, security
from models.schemas import (
    LandCreate, LandUpdate, LandResponse, Land,
    LandSectionCreate, LandSection,
    SectionDefinition, MessageResponse
)
from pydantic import BaseModel
from utils.geocoding import geocode_if_needed, coordinates_are_valid

router = APIRouter(prefix="/lands", tags=["lands"])


def normalize_energy_key(energy_key: Optional[str]) -> Optional[str]:
    """
    Normalize energy_key to match valid database values in lu_energy_type table.
    Maps common variations to correct database values.
    
    Valid values: 'solar', 'wind', 'hydroelectric', 'biomass', 'geothermal'
    """
    if not energy_key:
        return None
    
    # Convert to lowercase for case-insensitive matching
    energy_key_lower = str(energy_key).lower().strip()
    
    # Mapping of common variations to database values
    energy_key_mapping = {
        'hydro': 'hydroelectric',
        'hydro electric': 'hydroelectric',
        'hydro-electric': 'hydroelectric',
        'hydroelectric': 'hydroelectric',  # Already correct
        'solar': 'solar',
        'wind': 'wind',
        'biomass': 'biomass',
        'geothermal': 'geothermal',
        'geo-thermal': 'geothermal',
        'geo thermal': 'geothermal'
    }
    
    # Return normalized value or original if not in mapping
    normalized = energy_key_mapping.get(energy_key_lower)
    if normalized:
        return normalized
    
    # If exact match (case-insensitive) with valid values, return lowercase
    valid_keys = ['solar', 'wind', 'hydroelectric', 'biomass', 'geothermal']
    if energy_key_lower in valid_keys:
        return energy_key_lower
    
    # Return original if no mapping found (will fail foreign key constraint, but at least we tried)
    print(f"[normalize_energy_key] Warning: Unknown energy_key '{energy_key}', returning as-is")
    return energy_key_lower


# Admin Dashboard
@router.get("/admin/projects", response_model=List[Dict[str, Any]])
async def get_admin_projects(
    status_filter: Optional[str] = Query(None, description="Filter by status: submitted, under_review, approved, published, etc."),
    current_user: dict = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Get all projects for admin review (admin only)."""
    
    # Query to get all submitted projects with landowner details, investor interest count, and reviewer assignment status
    query = text("""
        SELECT 
            l.land_id,
            l.title,
            l.location_text,
            l.land_type,
            l.energy_key,
            l.capacity_mw,
            l.price_per_mwh,
            l.area_acres,
            l.status,
            l.timeline_text,
            l.contract_term_years,
            l.developer_name,
            l.admin_notes,
            l.project_priority,
            l.project_due_date,
            l.created_at,
            l.updated_at,
            l.published_at,
            l.landowner_id,
            CASE 
                WHEN l.site_image IS NOT NULL THEN true
                ELSE false
            END as has_site_image,
            u.email as landowner_email,
            u.first_name,
            u.last_name,
            u.phone,
            COALESCE(COUNT(DISTINCT ii.interest_id), 0) as investor_interest_count,
            CASE 
                WHEN COUNT(DISTINCT t.task_id) FILTER (WHERE t.assigned_to IS NOT NULL) > 0 THEN true
                ELSE false
            END as has_reviewers
        FROM lands l
        LEFT JOIN "user" u ON l.landowner_id = u.user_id
        LEFT JOIN investor_interests ii ON l.land_id = ii.land_id
        LEFT JOIN tasks t ON l.land_id = t.land_id AND t.assigned_to IS NOT NULL
        WHERE 1=1
        AND l.status != 'draft'
    """ + (" AND l.status = :status_filter" if status_filter else "") + """
        GROUP BY l.land_id, l.title, l.location_text, l.land_type, l.energy_key,
                 l.capacity_mw, l.price_per_mwh, l.area_acres, l.status,
                 l.timeline_text, l.contract_term_years, l.developer_name,
                 l.admin_notes, l.project_priority, l.project_due_date,
                 l.created_at, l.updated_at, l.published_at, l.landowner_id,
                 l.site_image, u.email, u.first_name, u.last_name, u.phone
        ORDER BY 
            CASE l.status
                WHEN 'submitted' THEN 1
                WHEN 'under_review' THEN 2
                WHEN 'approved' THEN 3
                WHEN 'published' THEN 4
                WHEN 'rtb' THEN 5
                WHEN 'interest_locked' THEN 6
                ELSE 7
            END,
            l.created_at DESC
    """)
    
    params = {}
    if status_filter:
        params["status_filter"] = status_filter
    
    results = db.execute(query, params).fetchall()
    
    projects = []
    for row in results:
        landowner_name = f"{row.first_name or ''} {row.last_name or ''}".strip() or row.landowner_email
        
        # Build image URL for site image if available
        image_url = None
        has_site_image = row.has_site_image if hasattr(row, 'has_site_image') else False
        if has_site_image:
            image_url = f"/api/lands/{row.land_id}/site-image"
        
        # Check if reviewers are assigned
        has_reviewers = row.has_reviewers if hasattr(row, 'has_reviewers') else False
        
        # Determine display status: if no reviewers assigned and status is not published/approved/rejected, show as pending
        display_status = row.status or "unknown"
        if not has_reviewers and display_status not in ['published', 'approved', 'rejected', 'rtb', 'interest_locked']:
            display_status = "pending"
        
        project = {
            "id": str(row.land_id),
            "land_id": str(row.land_id),
            "landowner_id": str(row.landowner_id) if row.landowner_id else None,
            "landownerName": landowner_name,
            "landownerEmail": row.landowner_email,
            "landownerPhone": row.phone,
            "location": row.location_text or "Not specified",
            "location_text": row.location_text or "Not specified",
            "projectType": row.land_type or "Not specified",
            "land_type": row.land_type or None,
            "energy_key": row.energy_key or "Not specified",
            "energyType": row.energy_key or "Not specified",
            "capacity": f"{row.capacity_mw} MW" if row.capacity_mw else "Not specified",
            "capacity_mw": float(row.capacity_mw) if row.capacity_mw else None,
            "pricePerMWh": float(row.price_per_mwh) if row.price_per_mwh else 0,
            "price_per_mwh": float(row.price_per_mwh) if row.price_per_mwh else None,
            "areaAcres": float(row.area_acres) if row.area_acres else 0,
            "area_acres": float(row.area_acres) if row.area_acres else None,
            "status": display_status,
            "original_status": row.status or "unknown",  # Keep original status for reference
            "has_reviewers": has_reviewers,
            "timeline": row.timeline_text or "Not specified",
            "timeline_text": row.timeline_text or None,
            "contractTerm": f"{row.contract_term_years} years" if row.contract_term_years else "Not specified",
            "contract_term_years": row.contract_term_years,
            "developerName": row.developer_name or "Not specified",
            "developer_name": row.developer_name or None,
            "adminNotes": row.admin_notes or "",
            "admin_notes": row.admin_notes or None,
            "description": row.admin_notes or None,
            "project_priority": row.project_priority,
            "project_due_date": row.project_due_date.isoformat() if row.project_due_date else None,
            "submittedDate": row.created_at.isoformat() if row.created_at else None,
            "created_at": row.created_at.isoformat() if row.created_at else None,
            "lastUpdated": row.updated_at.isoformat() if row.updated_at else None,
            "updated_at": row.updated_at.isoformat() if row.updated_at else None,
            "publishedAt": row.published_at.isoformat() if row.published_at else None,
            "title": row.title or f"{row.land_type} Project",
            "investorInterestCount": int(row.investor_interest_count) if row.investor_interest_count else 0,
            "has_site_image": has_site_image,
            "image_url": image_url,
            "first_name": row.first_name,
            "last_name": row.last_name,
            "phone": row.phone
        }
        projects.append(project)
    
    return projects


@router.get("/admin/summary", response_model=Dict[str, Any])
async def get_admin_summary(
    current_user: dict = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Get admin dashboard summary statistics (admin only)."""
    
    # Call stored procedure
    result = db.execute(text("SELECT * FROM get_admin_summary()")).fetchone()
    
    # Calculate pending reviews: include projects without reviewers
    pending_without_reviewers = result.pending_without_reviewers if hasattr(result, 'pending_without_reviewers') else 0
    # Pending reviews = submitted status + projects without reviewers (excluding those already counted as submitted)
    # The query already excludes projects with status 'submitted' from pending_without_reviewers, so we add them
    total_pending = result.pending_reviews + pending_without_reviewers
    
    return {
        "totalProjects": result.total_projects,
        "pendingReviews": total_pending,
        "underReview": result.under_review,
        "approved": result.approved,
        "published": result.published,
        "readyToBuy": result.ready_to_buy,
        "rejected": result.rejected,
        "totalLandArea": float(result.total_land_area),
        "totalCapacity": float(result.total_capacity),
        "estimatedRevenue": float(result.estimated_revenue) if hasattr(result, 'estimated_revenue') else 0.0,
        "totalLandowners": result.landowner_count,
        "totalInvestorInterests": int(result.total_interests) if result.total_interests else 0,
        "totalInvestors": int(result.total_investors) if result.total_investors else 0
    }

@router.get("/admin/deadline-alerts", response_model=List[Dict[str, Any]])
async def get_deadline_alerts(
    current_user: dict = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Get deadline alerts for projects and tasks (admin only).
    Returns alerts for projects and tasks with upcoming or overdue deadlines."""
    
    from datetime import datetime, timedelta, date as date_type
    
    now = datetime.now()
    alerts = []
    
    # Get project deadline alerts
    project_query = text("""
        SELECT 
            l.land_id,
            l.title,
            l.project_due_date,
            l.status,
            l.energy_key,
            u.first_name || ' ' || u.last_name as landowner_name,
            u.email as landowner_email
        FROM lands l
        LEFT JOIN "user" u ON l.landowner_id = u.user_id
        WHERE l.project_due_date IS NOT NULL
        AND l.status IN ('submitted', 'under_review', 'approved')
        AND l.project_due_date <= (CURRENT_DATE + INTERVAL '7 days')
        ORDER BY l.project_due_date ASC
    """)
    
    project_results = db.execute(project_query).fetchall()
    
    for row in project_results:
        due_date = row.project_due_date
        # Convert date to datetime for comparison
        if isinstance(due_date, str):
            due_date = datetime.fromisoformat(due_date.replace('Z', '+00:00'))
        elif isinstance(due_date, date_type):
            # It's a date object, convert to datetime at midnight
            due_date = datetime.combine(due_date, datetime.min.time())
        elif not isinstance(due_date, datetime):
            # Fallback: try to convert if it has date() method
            due_date = datetime.combine(due_date, datetime.min.time())
        
        diff = (due_date - now).total_seconds() / 3600  # hours
        diff_days = diff / 24
        
        energy_type = row.energy_key or "Project"
        landowner_name = row.landowner_name or row.landowner_email or "Unassigned"
        
        # Determine urgency
        if diff < 0:
            urgency = "critical"
            description = f"{row.title} - {energy_type} project is overdue"
        elif diff <= 24:
            urgency = "critical"
            description = f"{row.title} - Due in less than 24 hours"
        elif diff_days <= 3:
            urgency = "warning"
            description = f"{row.title} - Due in {int(diff_days)} days"
        else:
            urgency = "info"
            description = f"{row.title} - Due in {int(diff_days)} days"
        
        alerts.append({
            "id": f"project-{row.land_id}",
            "projectId": str(row.land_id),
            "landId": str(row.land_id),
            "taskId": None,
            "taskTitle": row.title,
            "projectTitle": row.title,
            "deadline": due_date.isoformat() if isinstance(due_date, datetime) else str(due_date),
            "description": description,
            "urgency": urgency,
            "projectName": row.title,
            "assignedTo": landowner_name,
            "type": "project"
        })
    
    # Get task deadline alerts
    task_query = text("""
        SELECT 
            t.task_id,
            t.land_id,
            t.title as task_title,
            t.description,
            t.due_date,
            t.status,
            t.assigned_to,
            l.title as project_title,
            l.energy_key,
            u1.first_name || ' ' || u1.last_name as assigned_to_name,
            u1.email as assigned_to_email
        FROM tasks t
        JOIN lands l ON t.land_id = l.land_id
        LEFT JOIN "user" u1 ON t.assigned_to = u1.user_id
        WHERE t.due_date IS NOT NULL
        AND t.status NOT IN ('completed', 'cancelled')
        AND t.due_date <= (CURRENT_DATE + INTERVAL '7 days')
        ORDER BY t.due_date ASC
    """)
    
    task_results = db.execute(task_query).fetchall()
    
    for row in task_results:
        due_date = row.due_date
        # Convert date to datetime for comparison
        if isinstance(due_date, str):
            due_date = datetime.fromisoformat(due_date.replace('Z', '+00:00'))
        elif isinstance(due_date, date_type):
            # It's a date object, convert to datetime at midnight
            due_date = datetime.combine(due_date, datetime.min.time())
        elif not isinstance(due_date, datetime):
            # Fallback: try to convert if it has date() method
            due_date = datetime.combine(due_date, datetime.min.time())
        
        diff = (due_date - now).total_seconds() / 3600  # hours
        diff_days = diff / 24
        
        assigned_to_name = row.assigned_to_name or row.assigned_to_email or "Unassigned"
        
        # Determine urgency
        if diff < 0:
            urgency = "critical"
            description = row.description or f"{row.task_title} - Task is overdue"
        elif diff <= 24:
            urgency = "critical"
            description = row.description or f"{row.task_title} - Due in less than 24 hours"
        elif diff_days <= 3:
            urgency = "warning"
            description = row.description or f"{row.task_title} - Due in {int(diff_days)} days"
        else:
            urgency = "info"
            description = row.description or f"{row.task_title} - Due in {int(diff_days)} days"
        
        alerts.append({
            "id": f"task-{row.task_id}",
            "projectId": str(row.land_id),
            "landId": str(row.land_id),
            "taskId": str(row.task_id),
            "taskTitle": row.task_title,
            "projectTitle": row.project_title,
            "deadline": due_date.isoformat() if isinstance(due_date, datetime) else str(due_date),
            "description": description,
            "urgency": urgency,
            "projectName": row.project_title,
            "assignedTo": assigned_to_name,
            "type": "task"
        })
    
    # Sort by urgency and deadline
    urgency_order = {"critical": 0, "warning": 1, "info": 2}
    alerts.sort(key=lambda x: (
        urgency_order.get(x["urgency"], 3),
        datetime.fromisoformat(x["deadline"].replace('Z', '+00:00')) if isinstance(x["deadline"], str) else x["deadline"]
    ))
    
    return alerts

@router.get("/admin/projects/{project_id}/details-with-tasks", response_model=Dict[str, Any])
async def get_project_details_with_tasks(
    project_id: UUID,
    current_user: dict = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Get project details with existing tasks for reviewer assignment (admin only)."""
   
    # Get project details using stored procedure
    project_result = db.execute(
        text("SELECT * FROM get_project_details_with_tasks(CAST(:project_id AS uuid))"),
        {"project_id": str(project_id)}
    ).fetchone()
   
    if not project_result:
        raise HTTPException(status_code=404, detail="Project not found")
   
    # Get existing tasks for this project using stored procedure
    tasks_result = db.execute(
        text("SELECT * FROM get_project_tasks(CAST(:project_id AS uuid))"),
        {"project_id": str(project_id)}
    ).fetchall()
   
    # Format project data
    landowner_name = f"{project_result.first_name or ''} {project_result.last_name or ''}".strip() or project_result.landowner_email
   
    project_data = {
        "id": str(project_result.land_id),
        "title": project_result.title,
        "location_text": project_result.location_text,
        "land_type": project_result.land_type,
        "energy_key": project_result.energy_key,
        "capacity_mw": project_result.capacity_mw,
        "price_per_mwh": float(project_result.price_per_mwh) if project_result.price_per_mwh else 0,
        "area_acres": float(project_result.area_acres) if project_result.area_acres else 0,
        "status": project_result.status,
        "timeline_text": project_result.timeline_text,
        "contract_term_years": project_result.contract_term_years,
        "developer_name": project_result.developer_name,
        "admin_notes": project_result.admin_notes,
        "project_priority": project_result.project_priority,
        "project_due_date": project_result.project_due_date.isoformat() if project_result.project_due_date else None,
        "created_at": project_result.created_at.isoformat() if project_result.created_at else None,
        "updated_at": project_result.updated_at.isoformat() if project_result.updated_at else None,
        "published_at": project_result.published_at.isoformat() if project_result.published_at else None,
        "landownerName": landowner_name,
        "landownerEmail": project_result.landowner_email,
        "landownerPhone": project_result.phone,
        "tasks": []
    }
   
    # Format tasks data
    for task in tasks_result:
        task_data = {
            "task_id": str(task.task_id),
            "land_id": str(task.land_id),
            "task_type": task.task_type,
            "description": task.description,
            "assigned_to": str(task.assigned_to) if task.assigned_to else None,
            "assigned_role": task.assigned_role,
            "status": task.status,
            "priority": task.priority,
            "due_date": task.due_date.isoformat() if task.due_date else None,
            "created_at": task.created_at.isoformat() if task.created_at else None,
            "updated_at": task.updated_at.isoformat() if task.updated_at else None,
            "assigned_to_name": task.assigned_to_name
        }
        project_data["tasks"].append(task_data)
   
    return project_data
 
@router.get("/admin/investor-interests", response_model=List[Dict[str, Any]])
async def get_admin_investor_interests(
    current_user: dict = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Get all investor interests with detailed information (admin only)."""
    
    # Call stored procedure
    results = db.execute(text("SELECT * FROM get_admin_investor_interests()")).fetchall()
    
    interests = []
    for row in results:
        interest = {
            "interestId": str(row.interest_id),
            "landId": str(row.land_id),
            "investorId": str(row.investor_id),
            "status": row.status,
            "comments": row.comments,
            "investmentAmount": float(row.investment_amount) if row.investment_amount else None,
            "createdAt": row.created_at.isoformat() if row.created_at else None,
            "updatedAt": row.updated_at.isoformat() if row.updated_at else None,
            "projectTitle": row.project_title,
            "projectLocation": row.project_location,
            "projectType": row.project_type,
            "projectCapacity": float(row.capacity_mw) if row.capacity_mw else None,
            "projectPrice": float(row.price_per_mwh) if row.price_per_mwh else None,
            "projectStatus": row.project_status,
            "investorName": f"{row.investor_first_name} {row.investor_last_name}".strip(),
            "investorEmail": row.investor_email,
            "investorPhone": row.investor_phone,
            "landownerName": row.landowner_name
        }
        interests.append(interest)
    
    return interests


# Landowner Dashboard
@router.get("/dashboard/summary", response_model=Dict[str, Any])
async def get_landowner_dashboard_summary(
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get landowner dashboard summary with statistics."""
    user_id = current_user["user_id"]
    
    # Call stored procedure
    result = db.execute(
        text("SELECT * FROM get_landowner_dashboard_summary(CAST(:user_id AS uuid))"),
        {"user_id": str(user_id)}
    ).fetchone()
    
    return {
        "totalLandArea": float(result.total_land_area) if result.total_land_area else 0,
        "draftProjects": int(result.draft_projects),
        "completedSubmissions": int(result.completed_submissions),
        "estimatedRevenue": float(result.estimated_revenue) if result.estimated_revenue else 0,
        "totalProjects": int(result.total_projects)
    }


@router.get("/dashboard/projects", response_model=List[Dict[str, Any]])
async def get_landowner_dashboard_projects(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    status_filter: Optional[str] = Query(None, description="Filter by status"),
    search: Optional[str] = Query(None, description="Search by name or location"),
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get landowner's projects for dashboard display."""
    user_id = current_user["user_id"]
    
    # Call stored procedure with parameters
    # SQLAlchemy doesn't support ::type syntax, so we use CAST
    # For NULL values, we pass them directly without casting
    results = db.execute(
        text("""
            SELECT * FROM get_landowner_dashboard_projects(
                CAST(:user_id AS uuid),
                :status_filter,
                :search,
                CAST(:skip AS integer),
                CAST(:limit AS integer)
            )
        """),
        {
            "user_id": str(user_id),
            "status_filter": status_filter,
            "search": search,
            "skip": skip,
            "limit": limit
        }
    ).fetchall()
    
    # Format projects using exact database column names (no duplicates)
    projects = []
    for row in results:
        # Calculate estimated revenue if price data exists
        estimated_revenue = 0
        if row.capacity_mw and row.price_per_mwh:
            # Annual revenue in millions (capacity_mw * price_per_mwh * 8760 hours / 1,000,000)
            estimated_revenue = float(row.capacity_mw) * float(row.price_per_mwh) * 8760 / 1000000
        
        project = {
            "land_id": str(row.land_id),
            "title": row.title or "Untitled Project",
            "location_text": row.location_text or "Location not specified",
            "energy_key": row.energy_key or "solar",
            "capacity_mw": float(row.capacity_mw) if row.capacity_mw else 0,
            "area_acres": float(row.area_acres) if row.area_acres else None,
            "status": row.status,
            "updated_at": row.updated_at.isoformat() if row.updated_at else None,
            "timeline_text": row.timeline_text or "Not specified",
            "price_per_mwh": float(row.price_per_mwh) if row.price_per_mwh else None,
            "estimated_revenue": round(estimated_revenue, 2),
            "description": row.description,
            "created_at": row.created_at.isoformat() if row.created_at else None
        }
        projects.append(project)
    
    return projects


# Land CRUD operations
@router.post("/", response_model=Land)
async def create_land(
    land_data: LandCreate,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create a new land entry (authenticated users)."""
    try:
        # Generate new land ID
        import uuid
        land_id = uuid.uuid4()
        
        # Geocode postcode if coordinates are missing or invalid
        coordinates_to_use = land_data.coordinates
        if land_data.post_code:
            geocoded_coords = await geocode_if_needed(
                postcode=land_data.post_code,
                existing_coordinates=land_data.coordinates
            )
            if geocoded_coords:
                coordinates_to_use = geocoded_coords
                print(f"[create_land] Geocoded postcode {land_data.post_code} to coordinates: {geocoded_coords}")
        
        # Prepare coordinates as JSON string if it's a dict
        coordinates_json = coordinates_to_use
        if isinstance(coordinates_json, dict):
            coordinates_json = json.dumps(coordinates_json)
        
        # Map energy_key from API to database
        energy_key_value = None
        if land_data.energy_key:
            # Convert to string if it's an enum
            energy_key_str = str(land_data.energy_key) if hasattr(land_data.energy_key, 'value') else land_data.energy_key
            # Normalize energy_key to match database values
            energy_key_value = normalize_energy_key(energy_key_str)
        
        # Ensure 'draft' status exists in lu_status table
        db.execute(text("""
            INSERT INTO lu_status(status_key, scope) 
            VALUES ('draft', 'land')
            ON CONFLICT (status_key) DO NOTHING
        """))
        
        # Create land with direct INSERT
        insert_query = text("""
            INSERT INTO lands (
                land_id, landowner_id, title, location_text, post_code, coordinates, 
                area_acres, land_type, energy_key, potential_partners, project_description
            ) VALUES (
                :land_id, :landowner_id, :title, :location_text, :post_code, :coordinates,
                :area_acres, :land_type, :energy_key, :potential_partners, :project_description
            )
        """)
        
        db.execute(insert_query, {
            "land_id": str(land_id),
            "landowner_id": current_user["user_id"],
            "title": land_data.title,
            "location_text": land_data.location_text,
            "post_code": land_data.post_code,
            "coordinates": coordinates_json,
            "area_acres": float(land_data.area_acres) if land_data.area_acres else None,
            "land_type": land_data.land_type,
            "energy_key": energy_key_value,
            "potential_partners": land_data.potential_partners,
            "project_description": land_data.project_description
        })
        
        db.commit()
        
        # Fetch the created land
        return await get_land(land_id, current_user, db)
        
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to create land: {str(e)}"
        )

@router.get("/", response_model=List[Land])
async def list_lands(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    status_filter: Optional[str] = Query(None, description="Filter by status"),
    owner_id: Optional[UUID] = Query(None, description="Filter by owner"),
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """List lands with optional filters."""
    # Build dynamic query based on user role and filters
    base_query = """
        SELECT l.land_id, l.landowner_id, l.title, l.location_text,
               l.coordinates, l.area_acres, l.land_type, l.status,
               l.admin_notes, l.energy_key, l.capacity_mw, l.price_per_mwh,
               l.timeline_text, l.contract_term_years, l.developer_name,
               l.published_at, l.interest_locked_at, l.created_at, l.updated_at
        FROM lands l
        WHERE 1=1
    """
    
    params = {"skip": skip, "limit": limit}
    
    # Apply filters based on user role
    user_roles = current_user.get("roles", [])
    if "administrator" not in user_roles:
        # Non-admin users can only see their own lands or published lands
        base_query += " AND (l.landowner_id = :current_user_id OR l.status = 'published')"
        params["current_user_id"] = current_user["user_id"]
    
    if status_filter:
        base_query += " AND l.status = :status_filter"
        params["status_filter"] = status_filter
    
    if owner_id:
        base_query += " AND l.landowner_id = :owner_id"
        params["owner_id"] = str(owner_id)
    
    base_query += " ORDER BY l.created_at DESC OFFSET :skip LIMIT :limit"
    
    results = db.execute(text(base_query), params).fetchall()
    
    return [
        Land(
            land_id=row.land_id,
            landowner_id=row.landowner_id,
            title=row.title,
            location_text=row.location_text,
            coordinates=row.coordinates,
            area_acres=Decimal(str(row.area_acres)) if row.area_acres else None,
            land_type=row.land_type,
            status=row.status,
            admin_notes=row.admin_notes,
            energy_key=row.energy_key,
            capacity_mw=Decimal(str(row.capacity_mw)) if row.capacity_mw else None,
            price_per_mwh=Decimal(str(row.price_per_mwh)) if row.price_per_mwh else None,
            timeline_text=row.timeline_text,
            contract_term_years=row.contract_term_years,
            developer_name=row.developer_name,
            published_at=row.published_at,
            interest_locked_at=row.interest_locked_at,
            created_at=row.created_at,
            updated_at=row.updated_at
        )
        for row in results
    ]

@router.get("/{land_id}", response_model=Land)
async def get_land(
    land_id: UUID,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get land by ID."""
    query = text("""
        SELECT l.land_id, l.landowner_id, l.title, l.location_text, l.post_code,
               l.coordinates, 
               l.area_acres,
               NULLIF(NULLIF(l.land_type, ''), ' ') as land_type,
               l.status,
               l.admin_notes, l.energy_key, l.capacity_mw, l.price_per_mwh,
               l.timeline_text, 
               l.contract_term_years,
               l.developer_name, l.potential_partners, l.project_description,
               l.project_priority, l.project_due_date,
               l.published_at, l.interest_locked_at, l.created_at, l.updated_at
        FROM lands l
        WHERE l.land_id = :land_id
    """)
    
    result = db.execute(query, {"land_id": str(land_id)}).fetchone()
    
    if not result:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Land not found"
        )
    
    # Debug: Log raw database values
    print(f"[get_land] Raw DB values for {land_id}:")
    print(f"  area_acres: {result.area_acres} (type: {type(result.area_acres)})")
    print(f"  land_type: '{result.land_type}' (type: {type(result.land_type)})")
    print(f"  contract_term_years: {result.contract_term_years} (type: {type(result.contract_term_years)})")
    
    # Check permissions - compare both as strings to avoid UUID/string mismatch
    user_roles = current_user.get("roles", [])
    user_id_str = str(current_user["user_id"])
    landowner_id_str = str(result.landowner_id)
    
    is_admin = "administrator" in user_roles
    is_owner = user_id_str == landowner_id_str
    is_published = result.status == "published"
    
    # Check if user is a reviewer assigned to this land
    is_reviewer = False
    if not (is_admin or is_owner or is_published):
        reviewer_check = text("""
            SELECT assigned_role FROM tasks 
            WHERE land_id = :land_id AND assigned_to = :user_id
            LIMIT 1
        """)
        reviewer_result = db.execute(
            reviewer_check, 
            {"land_id": str(land_id), "user_id": user_id_str}
        ).fetchone()
        
        if reviewer_result:
            is_reviewer = True
    
    # Check if user is an investor who has expressed interest in this land
    is_investor_with_interest = False
    if not (is_admin or is_owner or is_published or is_reviewer) and "investor" in user_roles:
        interest_check = text("""
            SELECT interest_id FROM investor_interests 
            WHERE land_id = :land_id AND investor_id = :user_id
            AND status IN ('pending', 'approved')
            AND (withdrawal_requested = FALSE OR withdrawal_status IS NULL OR withdrawal_status != 'approved')
            LIMIT 1
        """)
        interest_result = db.execute(
            interest_check, 
            {"land_id": str(land_id), "user_id": user_id_str}
        ).fetchone()
        
        if interest_result:
            is_investor_with_interest = True
    
    if not (is_admin or is_owner or is_published or is_reviewer or is_investor_with_interest):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions to view this land"
        )
    
    # Handle area_acres - convert to Decimal if not None, handle empty strings and 0
    area_acres_value = None
    if result.area_acres is not None:
        try:
            # Convert to string first to handle Decimal types from DB
            area_str = str(result.area_acres).strip()
            if area_str and area_str != '' and area_str.lower() != 'none':
                area_acres_value = Decimal(area_str)
        except (ValueError, TypeError) as e:
            print(f"[get_land] Error converting area_acres: {e}")
            area_acres_value = None
    
    # Handle land_type - convert empty strings to None
    land_type_value = None
    if result.land_type:
        land_type_str = str(result.land_type).strip()
        if land_type_str and land_type_str != '' and land_type_str.lower() != 'none':
            land_type_value = land_type_str
    
    # Handle contract_term_years - convert to int if not None
    contract_term_value = None
    if result.contract_term_years is not None:
        try:
            # Convert to string first to handle Decimal/numeric types from DB
            term_str = str(result.contract_term_years).strip()
            if term_str and term_str != '' and term_str.lower() != 'none':
                contract_term_value = int(float(term_str))  # Handle both int and float
        except (ValueError, TypeError) as e:
            print(f"[get_land] Error converting contract_term_years: {e}")
            contract_term_value = None
    
    # Debug: Log processed values
    print(f"[get_land] Processed values:")
    print(f"  area_acres_value: {area_acres_value}")
    print(f"  land_type_value: {land_type_value}")
    print(f"  contract_term_value: {contract_term_value}")
    
    return Land(
        land_id=result.land_id,
        landowner_id=result.landowner_id,
        title=result.title,
        location_text=result.location_text,
        post_code=result.post_code,
        coordinates=result.coordinates,
        area_acres=area_acres_value,
        land_type=land_type_value,
        status=result.status,
        admin_notes=result.admin_notes,
        energy_key=result.energy_key,
        capacity_mw=Decimal(str(result.capacity_mw)) if result.capacity_mw else None,
        price_per_mwh=Decimal(str(result.price_per_mwh)) if result.price_per_mwh else None,
        timeline_text=result.timeline_text,
        contract_term_years=contract_term_value,
        developer_name=result.developer_name,
        potential_partners=result.potential_partners,
        project_description=result.project_description,
        project_priority=result.project_priority,
        project_due_date=result.project_due_date,
        published_at=result.published_at,
        interest_locked_at=result.interest_locked_at,
        created_at=result.created_at,
        updated_at=result.updated_at
    )

@router.put("/{land_id}", response_model=LandResponse)
async def update_land(
    land_id: UUID,
    land_update: LandUpdate,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Update land information (owner or admin only)."""
    # Check if land exists and user has permission
    land_check = text("""
        SELECT landowner_id, status FROM lands WHERE land_id = :land_id
    """)
    
    land_result = db.execute(land_check, {"land_id": str(land_id)}).fetchone()
    
    if not land_result:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Land not found"
        )
    
    # Check permissions - convert both to strings for comparison
    user_roles = current_user.get("roles", [])
    user_id_str = str(current_user["user_id"])
    owner_id_str = str(land_result.landowner_id)
    
    is_admin = "administrator" in user_roles
    is_owner = owner_id_str == user_id_str
    
    if not (is_admin or is_owner):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions to update this land"
        )
    
    # Prevent reverting status from submitted (or any non-draft status) back to draft
    # Once a project is submitted, it cannot be changed back to draft
    current_status = land_result.status
    update_data = land_update.dict(exclude_unset=True)
    
    if 'status' in update_data and update_data['status'] == 'draft':
        if current_status != 'draft':
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Cannot revert project status from '{current_status}' to 'draft'. Once submitted, a project cannot revert to draft status."
            )
    
    # Geocode postcode if it's being updated or if coordinates are missing/invalid
    postcode_to_geocode = None
    existing_coords = None
    
    # Check if postcode is being updated
    if 'post_code' in update_data and update_data['post_code']:
        postcode_to_geocode = update_data['post_code']
        # Get existing coordinates from database if not being updated
        if 'coordinates' not in update_data:
            existing_coords_query = text("SELECT coordinates FROM lands WHERE land_id = :land_id")
            existing_coords_result = db.execute(existing_coords_query, {"land_id": str(land_id)}).fetchone()
            if existing_coords_result and existing_coords_result.coordinates:
                existing_coords = existing_coords_result.coordinates
                if isinstance(existing_coords, str):
                    try:
                        existing_coords = json.loads(existing_coords)
                    except:
                        existing_coords = None
        else:
            existing_coords = update_data.get('coordinates')
    
    # Perform geocoding if needed
    if postcode_to_geocode:
        geocoded_coords = await geocode_if_needed(
            postcode=postcode_to_geocode,
            existing_coordinates=existing_coords
        )
        if geocoded_coords:
            # Update coordinates in the update_data if they weren't explicitly set
            if 'coordinates' not in update_data or not coordinates_are_valid(update_data.get('coordinates')):
                update_data['coordinates'] = geocoded_coords
                print(f"[update_land] Geocoded postcode {postcode_to_geocode} to coordinates: {geocoded_coords}")
    
    # Build dynamic update query
    update_fields = []
    params = {"land_id": str(land_id)}
    
    # Debug: Log incoming update data
    print(f"[update_land] Received update data for land_id {land_id}:")
    print(f"[update_land] Update data keys: {list(update_data.keys())}")
    for key, value in update_data.items():
        print(f"[update_land]   {key}: {value} (type: {type(value)})")
    
    for field, value in update_data.items():
        # Map energy_key field (no column mapping needed - matches database)
        db_field = field
        
        # String fields
        if field in ["title", "location_text", "post_code", "land_type", "admin_notes", "energy_key", 
                     "timeline_text", "developer_name", "project_priority", "potential_partners", "project_description"]:
            update_fields.append(f"{db_field} = :{field}")
            # Convert enum to string if needed and normalize energy_key
            if field == "energy_key" and value:
                energy_key_str = str(value) if hasattr(value, 'value') else value
                # Normalize energy_key to match database values
                params[field] = normalize_energy_key(energy_key_str)
            else:
                # Handle None/null values - set to NULL in database
                params[field] = value if value is not None else None
        # Numeric fields
        elif field in ["area_acres", "capacity_mw", "price_per_mwh"]:
            update_fields.append(f"{field} = :{field}")
            params[field] = float(value) if value is not None else None
        # Integer fields
        elif field in ["contract_term_years"]:
            update_fields.append(f"{field} = :{field}")
            params[field] = int(value) if value is not None else None
        # JSON fields
        elif field == "coordinates":
            # Use explicit JSONB cast in SQL for safety
            update_fields.append(f"{field} = CAST(:{field} AS JSONB)")
            if value is not None:
                # Normalize coordinates - prefer lat/lng format, remove duplicates
                if isinstance(value, dict):
                    # Normalize to use only lat/lng (not latitude/longitude)
                    normalized_coords = {}
                    if 'lat' in value or 'latitude' in value:
                        normalized_coords['lat'] = value.get('lat') or value.get('latitude')
                    if 'lng' in value or 'longitude' in value:
                        normalized_coords['lng'] = value.get('lng') or value.get('longitude')
                    # Only include if we have both lat and lng
                    if 'lat' in normalized_coords and 'lng' in normalized_coords:
                        params[field] = json.dumps(normalized_coords)
                    else:
                        params[field] = json.dumps(value)
                elif isinstance(value, str):
                    # If already a JSON string, try to parse and normalize
                    try:
                        parsed = json.loads(value)
                        if isinstance(parsed, dict):
                            normalized_coords = {}
                            if 'lat' in parsed or 'latitude' in parsed:
                                normalized_coords['lat'] = parsed.get('lat') or parsed.get('latitude')
                            if 'lng' in parsed or 'longitude' in parsed:
                                normalized_coords['lng'] = parsed.get('lng') or parsed.get('longitude')
                            if 'lat' in normalized_coords and 'lng' in normalized_coords:
                                params[field] = json.dumps(normalized_coords)
                            else:
                                params[field] = value
                        else:
                            params[field] = value
                    except:
                        params[field] = value
                else:
                    # Try to convert to JSON string
                    params[field] = json.dumps(value)
            else:
                params[field] = None
        # Date/time fields
        elif field == "project_due_date":
            update_fields.append(f"{field} = :{field}")
            # Convert string to timestamp if needed
            params[field] = value
        else:
            # Log unhandled fields
            print(f"[update_land] WARNING: Unhandled field '{field}' with value '{value}'")
    
    # Debug: Log what will be updated
    print(f"[update_land] Fields to update: {update_fields}")
    print(f"[update_land] Parameters: {params}")
    
    try:
        if update_fields:
            update_query = text(f"""
                UPDATE lands 
                SET {', '.join(update_fields)}, updated_at = CURRENT_TIMESTAMP
                WHERE land_id = :land_id
            """)
            
            print(f"[update_land] Executing query: {update_query}")
            db.execute(update_query, params)
            db.commit()
            print(f"[update_land] Update successful")
        else:
            print(f"[update_land] No fields to update")
        
        return await get_land(land_id, current_user, db)
    
    except Exception as e:
        db.rollback()
        error_msg = str(e)
        print(f"[update_land] Error updating land: {error_msg}")
        print(f"[update_land] Error type: {type(e).__name__}")
        import traceback
        print(f"[update_land] Traceback: {traceback.format_exc()}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to update land: {error_msg}"
        )

@router.post("/{land_id}/toggle-publish", response_model=Dict[str, Any])
async def toggle_publish_land(
    land_id: UUID,
    current_user: dict = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Toggle publish status of a land (admin only)."""
    # Check if land exists - get status from lands table
    land_check = text("""
        SELECT land_id, status, published_at 
        FROM lands 
        WHERE land_id = :land_id
    """)
    
    land_result = db.execute(land_check, {"land_id": str(land_id)}).fetchone()
    
    if not land_result:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Land not found"
        )
    
    # Toggle publish status
    # If currently published, change to under_review; otherwise publish it
    current_status = land_result.status
    new_status = "under_review" if current_status == "published" else "published"
    published_at = datetime.utcnow() if new_status == "published" else None
    
    # Update the land status
    update_query = text("""
        UPDATE lands 
        SET status = :new_status, 
            published_at = :published_at,
            updated_at = CURRENT_TIMESTAMP
        WHERE land_id = :land_id
        RETURNING land_id, status, published_at, updated_at
    """)
    
    result = db.execute(update_query, {
        "land_id": str(land_id),
        "new_status": new_status,
        "published_at": published_at
    }).fetchone()
    
    if not result:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to update land status"
        )
    
    db.commit()
    
    return {
        "land_id": str(result.land_id),
        "status": result.status,
        "published_at": result.published_at.isoformat() if result.published_at else None,
        "updated_at": result.updated_at.isoformat() if result.updated_at else None,
        "message": f"Land {new_status} successfully"
    }

@router.delete("/{land_id}", response_model=MessageResponse)
async def delete_land(
    land_id: UUID,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Delete land (owner or admin only)."""
    # Check if land exists and user has permission
    land_check = text("""
        SELECT landowner_id, status FROM lands WHERE land_id = :land_id
    """)
    
    land_result = db.execute(land_check, {"land_id": str(land_id)}).fetchone()
    
    if not land_result:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Land not found"
        )
    
    # Check permissions - convert both to strings for comparison
    user_roles = current_user.get("roles", [])
    user_id_str = str(current_user["user_id"])
    owner_id_str = str(land_result.landowner_id)
    
    is_admin = "administrator" in user_roles
    is_owner = owner_id_str == user_id_str
    
    if not (is_admin or is_owner):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions to delete this land"
        )
    
    # Only allow deletion of draft lands
    if land_result.status != "draft":
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only draft lands can be deleted"
        )
    
    try:
        # Delete in order to avoid foreign key constraint issues with triggers
        # The trigger log_document_change() fires AFTER DELETE on documents and tries to insert
        # into document_audit_trail, but the document_id foreign key constraint fails.
        # We need to delete audit trail entries first, then temporarily disable the trigger.
        
        # Step 1: Delete all document_audit_trail entries for this land
        delete_audit_query = text("""
            DELETE FROM document_audit_trail 
            WHERE land_id = :land_id
               OR document_id IN (
                   SELECT document_id FROM documents WHERE land_id = :land_id
               )
        """)
        db.execute(delete_audit_query, {"land_id": str(land_id)})
        
        # Step 2: Temporarily disable the trigger to avoid foreign key constraint issues
        disable_trigger_query = text("ALTER TABLE documents DISABLE TRIGGER trg_log_document_change")
        db.execute(disable_trigger_query)
        
        try:
            # Step 3: Delete documents (trigger is disabled, so no audit trail insertion)
            delete_documents_query = text("DELETE FROM documents WHERE land_id = :land_id")
            db.execute(delete_documents_query, {"land_id": str(land_id)})
            
            # Step 4: Delete the land (this will cascade delete land_sections, tasks, investor_interests, etc.)
            delete_land_query = text("DELETE FROM lands WHERE land_id = :land_id")
            result = db.execute(delete_land_query, {"land_id": str(land_id)})
            
            if result.rowcount == 0:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail="Land not found or already deleted"
                )
        finally:
            # Step 5: Re-enable the trigger
            enable_trigger_query = text("ALTER TABLE documents ENABLE TRIGGER trg_log_document_change")
            db.execute(enable_trigger_query)
        
        db.commit()
        return MessageResponse(message="Land deleted successfully")
        
    except HTTPException:
        db.rollback()
        # Make sure trigger is re-enabled even if there's an error
        try:
            db.execute(text("ALTER TABLE documents ENABLE TRIGGER trg_log_document_change"))
        except:
            pass
        raise
    except Exception as e:
        db.rollback()
        # Make sure trigger is re-enabled even if there's an error
        try:
            db.execute(text("ALTER TABLE documents ENABLE TRIGGER trg_log_document_change"))
        except:
            pass
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error deleting land: {str(e)}"
        )

# Land status management
@router.put("/{land_id}/submit", response_model=MessageResponse)
async def submit_land_for_review(
    land_id: UUID,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Submit land for review (owner only). Changes status from draft to under_review. Once submitted, cannot revert to draft."""
    # First check if land exists and belongs to user
    check_query = text("""
        SELECT landowner_id, status FROM lands WHERE land_id = :land_id
    """)
    
    land_result = db.execute(check_query, {"land_id": str(land_id)}).fetchone()
    
    if not land_result:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Land not found"
        )
    
    # Check permissions - convert both to strings for comparison
    user_id_str = str(current_user["user_id"])
    landowner_id_str = str(land_result.landowner_id)
    
    if landowner_id_str != user_id_str:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not authorized to submit this land"
        )
    
    # Prevent reverting from under_review (or any non-draft status) back to draft
    if land_result.status != 'draft':
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Can only submit draft lands. Current status: {land_result.status}. Once submitted, a project cannot revert to draft status."
        )
    
    # Update status from draft to under_review
    # This is a one-way transition - once submitted, it cannot be draft again
    update_query = text("""
        UPDATE lands 
        SET status = 'under_review', updated_at = CURRENT_TIMESTAMP
        WHERE land_id = :land_id 
        AND landowner_id = :owner_id 
        AND status = 'draft'
    """)
    
    try:
        result = db.execute(update_query, {
            "land_id": str(land_id),
            "owner_id": str(current_user["user_id"])
        })
        
        # Check if any rows were updated
        if result.rowcount == 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Failed to submit land. The project may have already been submitted or does not exist."
            )
        
        # Get project title for notification
        land_title_query = text("SELECT title FROM lands WHERE land_id = :land_id")
        land_title_result = db.execute(land_title_query, {"land_id": str(land_id)}).fetchone()
        project_title = land_title_result.title if land_title_result else "Project"
        
        # Get admin users to notify
        admin_query = text("""
            SELECT DISTINCT u.user_id
            FROM "user" u
            JOIN user_roles ur ON u.user_id = ur.user_id
            WHERE ur.role_key = 'administrator' AND u.is_active = true
        """)
        admin_users = db.execute(admin_query).fetchall()
        admin_user_ids = [str(row.user_id) for row in admin_users]
        
        # Create notifications
        try:
            from utils.notifications import notify_project_submitted
            notify_project_submitted(
                db=db,
                land_id=str(land_id),
                project_title=project_title,
                landowner_id=str(current_user["user_id"]),
                admin_user_ids=admin_user_ids
            )
        except Exception as e:
            print(f"Error creating project submission notification: {str(e)}")
        
        db.commit()
        return MessageResponse(message="Land submitted for review successfully. Status changed to under_review.")
        
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error submitting land: {str(e)}"
        )

@router.post("/{land_id}/publish", response_model=MessageResponse)
async def publish_land(
    land_id: UUID,
    current_user: dict = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Publish land (admin only)."""
    query = text("SELECT sp_publish_land(:land_id) as success")
    
    try:
        result = db.execute(query, {"land_id": str(land_id)}).fetchone()
        db.commit()
        
        if result and result.success:
            return MessageResponse(message="Land published successfully")
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Failed to publish land"
            )
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Error publishing land: {str(e)}"
        )

@router.post("/{land_id}/mark-rtb", response_model=MessageResponse)
async def mark_land_ready_to_buy(
    land_id: UUID,
    current_user: dict = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Mark land as ready to buy (admin only)."""
    query = text("SELECT sp_land_mark_rtb(:land_id) as success")
    
    try:
        result = db.execute(query, {"land_id": str(land_id)}).fetchone()
        db.commit()
        
        if result and result.success:
            return MessageResponse(message="Land marked as ready to buy successfully")
        else:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Failed to mark land as ready to buy"
            )
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Error marking land as RTB: {str(e)}"
        )

@router.get("/marketplace/published", response_model=List[Dict[str, Any]])
async def get_published_lands(
    skip: int = 0,
    limit: int = 100,
    energy_type: Optional[str] = None,
    min_capacity: Optional[float] = None,
    max_capacity: Optional[float] = None,
    min_price: Optional[float] = None,
    max_price: Optional[float] = None,
    location: Optional[str] = None,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security),
    db: Session = Depends(get_db)
):
    """Get all published lands for marketplace (public endpoint).
    Excludes interest_locked projects unless the current user is the investor who expressed interest."""
    
    # Try to get current user if authenticated (optional)
    current_user_id = None
    try:
        if credentials:
            from auth import verify_token
            token = credentials.credentials
            payload = verify_token(token)
            if payload:
                user_id = payload.get("user_id")
                if user_id:
                    current_user_id = str(user_id)
    except Exception:
        # If authentication fails, continue as unauthenticated user
        pass
    
    # Build dynamic query for published lands
    base_query = """
        SELECT 
            l.land_id,
            l.title,
            l.location_text,
            l.coordinates,
            l.land_type,
            l.energy_key,
            l.capacity_mw,
            l.price_per_mwh,
            l.area_acres,
            l.timeline_text,
            l.contract_term_years,
            l.developer_name,
            l.published_at,
            l.created_at,
            l.status,
            CASE 
                WHEN l.site_image IS NOT NULL THEN true
                ELSE false
            END as has_site_image,
            u.first_name || ' ' || u.last_name as landowner_name,
            u.email as landowner_email,
            COUNT(DISTINCT CASE 
                WHEN ii.status IN ('pending', 'approved') 
                AND (ii.withdrawal_requested = FALSE OR ii.withdrawal_status IS NULL OR ii.withdrawal_status != 'approved')
                THEN ii.interest_id 
            END) as interest_count
        FROM lands l
        LEFT JOIN "user" u ON l.landowner_id = u.user_id
        LEFT JOIN investor_interests ii ON l.land_id = ii.land_id
        WHERE l.status = 'published'
    """
    
    params = {"skip": skip, "limit": limit}
    
    # Exclude ALL projects with active interests from marketplace
    # Projects with interests should only be visible in "My Interest" section, not marketplace
    base_query += """
        AND NOT EXISTS (
            SELECT 1 FROM investor_interests ii_active
            WHERE ii_active.land_id = l.land_id
            AND ii_active.status IN ('pending', 'approved')
            AND (ii_active.withdrawal_requested = FALSE 
                 OR ii_active.withdrawal_status IS NULL 
                 OR ii_active.withdrawal_status != 'approved')
        )
    """
    
    # Apply filters
    if energy_type:
        base_query += " AND l.energy_key = :energy_type"
        params["energy_type"] = energy_type
    
    if min_capacity is not None:
        base_query += " AND l.capacity_mw >= :min_capacity"
        params["min_capacity"] = min_capacity
    
    if max_capacity is not None:
        base_query += " AND l.capacity_mw <= :max_capacity"
        params["max_capacity"] = max_capacity
    
    if min_price is not None:
        base_query += " AND l.price_per_mwh >= :min_price"
        params["min_price"] = min_price
    
    if max_price is not None:
        base_query += " AND l.price_per_mwh <= :max_price"
        params["max_price"] = max_price
    
    if location:
        base_query += " AND l.location_text ILIKE :location"
        params["location"] = f"%{location}%"
    
    base_query += """
        GROUP BY l.land_id, l.title, l.location_text, l.coordinates, l.land_type,
                 l.energy_key, l.capacity_mw, l.price_per_mwh, l.area_acres,
                 l.timeline_text, l.contract_term_years, l.developer_name,
                 l.published_at, l.created_at, l.site_image, u.first_name, u.last_name, u.email
        ORDER BY l.published_at DESC NULLS LAST, l.created_at DESC
        OFFSET :skip LIMIT :limit
    """
    
    results = db.execute(text(base_query), params).fetchall()
    
    projects = []
    for row in results:
        # Build image URL - use site_image if available, otherwise null
        image_url = None
        has_site_image = False
        if hasattr(row, 'has_site_image') and row.has_site_image:
            image_url = f"/api/lands/{row.land_id}/site-image"
            has_site_image = True
        
        project = {
            "id": str(row.land_id),
            "land_id": str(row.land_id),
            "title": row.title or f"{row.energy_key or 'Energy'} Project",
            "name": row.title or f"{row.energy_key or 'Energy'} Project",
            "type": row.energy_key or "Not specified",
            "energyType": row.energy_key or "Not specified",
            "energy_key": row.energy_key or "Not specified",
            "location": row.location_text or "Not specified",
            "location_text": row.location_text or "Not specified",
            "coordinates": row.coordinates,
            "capacity": float(row.capacity_mw) if row.capacity_mw else 0,
            "capacityMW": float(row.capacity_mw) if row.capacity_mw else 0,
            "capacity_mw": float(row.capacity_mw) if row.capacity_mw else None,
            "price": float(row.price_per_mwh) if row.price_per_mwh else 0,
            "pricePerMWh": float(row.price_per_mwh) if row.price_per_mwh else 0,
            "price_per_mwh": float(row.price_per_mwh) if row.price_per_mwh else None,
            "areaAcres": float(row.area_acres) if row.area_acres else 0,
            "area_acres": float(row.area_acres) if row.area_acres else None,
            "timeline": row.timeline_text or "Not specified",
            "contract": f"{row.contract_term_years} years" if row.contract_term_years else "Not specified",
            "contractTerm": row.contract_term_years,
            "landType": row.land_type,
            "land_type": row.land_type,
            "developerName": row.developer_name,
            "landownerName": row.landowner_name or row.landowner_email or "Unknown",
            "publishedAt": row.published_at.isoformat() if row.published_at else None,
            "published_at": row.published_at.isoformat() if row.published_at else None,
            "published_date": row.published_at.isoformat() if row.published_at else None,
            "interestCount": row.interest_count or 0,
            "status": row.status or "published",
            "isAvailable": (row.interest_count or 0) == 0,  # Only available if no active interests
            "image_url": image_url,
            "has_site_image": has_site_image
        }
        projects.append(project)
    
    return projects

@router.get("/admin/marketplace/published", response_model=List[Dict[str, Any]])
async def get_admin_marketplace_projects(
    skip: int = 0,
    limit: int = 1000,  # Increased default limit to ensure all projects are returned
    energy_type: Optional[str] = None,
    min_capacity: Optional[float] = None,
    max_capacity: Optional[float] = None,
    min_price: Optional[float] = None,
    max_price: Optional[float] = None,
    location: Optional[str] = None,
    current_user: dict = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Get all published lands for admin marketplace view.
    Includes ALL published projects regardless of interest status.
    Shows interest_locked badge for projects with active interests."""
    
    # Build dynamic query for published lands - fetch ALL published projects
    # Include projects with interests and calculate interest_count for each
    base_query = """
        SELECT 
            l.land_id,
            l.title,
            l.location_text,
            l.coordinates,
            l.land_type,
            l.energy_key,
            l.capacity_mw,
            l.price_per_mwh,
            l.area_acres,
            l.timeline_text,
            l.contract_term_years,
            l.developer_name,
            l.published_at,
            l.created_at,
            l.status,
            CASE 
                WHEN l.site_image IS NOT NULL THEN true
                ELSE false
            END as has_site_image,
            u.first_name || ' ' || u.last_name as landowner_name,
            u.email as landowner_email,
            COALESCE((
                SELECT COUNT(DISTINCT ii_sub.interest_id)
                FROM investor_interests ii_sub
                WHERE ii_sub.land_id = l.land_id
                AND ii_sub.status IN ('pending', 'approved')
                AND (ii_sub.withdrawal_requested = FALSE 
                     OR ii_sub.withdrawal_status IS NULL 
                     OR ii_sub.withdrawal_status != 'approved')
            ), 0) as interest_count
        FROM lands l
        LEFT JOIN "user" u ON l.landowner_id = u.user_id
        WHERE l.status = 'published'
    """
    
    params = {"skip": skip, "limit": limit}
    
    # Note: We do NOT exclude projects with interests for admin view
    # The query uses a subquery for interest_count, so no GROUP BY is needed
    
    # Apply filters
    if energy_type:
        base_query += " AND l.energy_key = :energy_type"
        params["energy_type"] = energy_type
    
    if min_capacity is not None:
        base_query += " AND l.capacity_mw >= :min_capacity"
        params["min_capacity"] = min_capacity
    
    if max_capacity is not None:
        base_query += " AND l.capacity_mw <= :max_capacity"
        params["max_capacity"] = max_capacity
    
    if min_price is not None:
        base_query += " AND l.price_per_mwh >= :min_price"
        params["min_price"] = min_price
    
    if max_price is not None:
        base_query += " AND l.price_per_mwh <= :max_price"
        params["max_price"] = max_price
    
    if location:
        base_query += " AND l.location_text ILIKE :location"
        params["location"] = f"%{location}%"
    
    # No GROUP BY needed since we're using a subquery for interest_count
    base_query += """
        ORDER BY l.published_at DESC NULLS LAST, l.created_at DESC
        OFFSET :skip LIMIT :limit
    """
    
    results = db.execute(text(base_query), params).fetchall()
    
    projects = []
    for row in results:
        # Build image URL - use site_image if available, otherwise null
        image_url = None
        has_site_image = False
        if hasattr(row, 'has_site_image') and row.has_site_image:
            image_url = f"/api/lands/{row.land_id}/site-image"
            has_site_image = True
        
        # Get interest_count - ensure it's an integer
        interest_count = int(row.interest_count) if row.interest_count is not None else 0
        is_interest_locked = interest_count > 0
        
        project = {
            "id": str(row.land_id),
            "land_id": str(row.land_id),
            "title": row.title or f"{row.energy_key or 'Energy'} Project",
            "name": row.title or f"{row.energy_key or 'Energy'} Project",
            "type": row.energy_key or "Not specified",
            "energyType": row.energy_key or "Not specified",
            "energy_key": row.energy_key or "Not specified",
            "location": row.location_text or "Not specified",
            "location_text": row.location_text or "Not specified",
            "coordinates": row.coordinates,
            "capacity": float(row.capacity_mw) if row.capacity_mw else 0,
            "capacityMW": float(row.capacity_mw) if row.capacity_mw else 0,
            "capacity_mw": float(row.capacity_mw) if row.capacity_mw else None,
            "price": float(row.price_per_mwh) if row.price_per_mwh else 0,
            "pricePerMWh": float(row.price_per_mwh) if row.price_per_mwh else 0,
            "price_per_mwh": float(row.price_per_mwh) if row.price_per_mwh else None,
            "areaAcres": float(row.area_acres) if row.area_acres else 0,
            "area_acres": float(row.area_acres) if row.area_acres else None,
            "timeline": row.timeline_text or "Not specified",
            "contract": f"{row.contract_term_years} years" if row.contract_term_years else "Not specified",
            "contractTerm": row.contract_term_years,
            "landType": row.land_type,
            "land_type": row.land_type,
            "developerName": row.developer_name,
            "landownerName": row.landowner_name or row.landowner_email or "Unknown",
            "publishedAt": row.published_at.isoformat() if row.published_at else None,
            "published_at": row.published_at.isoformat() if row.published_at else None,
            "published_date": row.published_at.isoformat() if row.published_at else None,
            "interestCount": interest_count,
            "interest_count": interest_count,
            "status": row.status or "published",
            "isAvailable": interest_count == 0,  # Available if no active interests
            "is_interest_locked": is_interest_locked,  # New field for admin view
            "image_url": image_url,
            "has_site_image": has_site_image
        }
        projects.append(project)
    
    return projects

# Land sections management
@router.get("/{land_id}/sections", response_model=List[LandSection])
async def get_land_sections(
    land_id: UUID,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get all sections for a land."""
    # First check if user can access this land
    await get_land(land_id, current_user, db)
    
    query = text("""
        SELECT ls.land_section_id, ls.land_id, ls.section_definition_id,
               ls.section_data, ls.created_at, ls.updated_at,
               sd.section_name, sd.section_type, sd.is_required
        FROM land_sections ls
        JOIN section_definitions sd ON ls.section_definition_id = sd.section_definition_id
        WHERE ls.land_id = :land_id
        ORDER BY sd.section_name
    """)
    
    results = db.execute(query, {"land_id": str(land_id)}).fetchall()
    
    return [
        LandSection(
            land_section_id=row.land_section_id,
            land_id=row.land_id,
            section_definition_id=row.section_definition_id,
            section_data=row.section_data,
            created_at=row.created_at,
            updated_at=row.updated_at
        )
        for row in results
    ]

@router.post("/{land_id}/sections", response_model=LandSection)
async def create_land_section(
    land_id: UUID,
    section_data: LandSectionCreate,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Create a new section for a land."""
    # Check if user owns the land or is admin
    land_check = text("SELECT owner_id FROM lands WHERE land_id = :land_id")
    land_result = db.execute(land_check, {"land_id": str(land_id)}).fetchone()
    
    if not land_result:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Land not found"
        )
    
    # Check permissions - convert both to strings for comparison
    user_roles = current_user.get("roles", [])
    user_id_str = str(current_user["user_id"])
    owner_id_str = str(land_result.owner_id)
    
    is_admin = "administrator" in user_roles
    is_owner = owner_id_str == user_id_str
    
    if not (is_admin or is_owner):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions"
        )
    
    # Use stored procedure to assign section
    query = text("""
        SELECT sp_assign_section(:land_id, :section_definition_id, :section_data) as section_id
    """)
    
    result = db.execute(query, {
        "land_id": str(land_id),
        "section_definition_id": str(section_data.section_definition_id),
        "section_data": section_data.section_data
    }).fetchone()
    
    db.commit()
    
    if not result or not result.section_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Failed to create section"
        )
    
    # Fetch the created section
    section_query = text("""
        SELECT land_section_id, land_id, section_definition_id, section_data, created_at, updated_at
        FROM land_sections
        WHERE land_section_id = :section_id
    """)
    
    section_result = db.execute(section_query, {"section_id": result.section_id}).fetchone()
    
    return LandSection(
        land_section_id=section_result.land_section_id,
        land_id=section_result.land_id,
        section_definition_id=section_result.section_definition_id,
        section_data=section_result.section_data,
        created_at=section_result.created_at,
        updated_at=section_result.updated_at
    )

@router.get("/sections/definitions", response_model=List[SectionDefinition])
async def get_section_definitions(
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get all available section definitions."""
    query = text("""
        SELECT section_definition_id, section_name, section_type, is_required, created_at
        FROM section_definitions
        ORDER BY section_name
    """)
    
    results = db.execute(query).fetchall()
    
    return [
        SectionDefinition(
            section_definition_id=row.section_definition_id,
            section_name=row.section_name,
            section_type=row.section_type,
            is_required=row.is_required,
            created_at=row.created_at
        )
        for row in results
    ]


# Pydantic models for document role mappings
class DocumentRoleMappingsRequest(BaseModel):
    mappings: Dict[str, List[str]]

# Project Document Role Mappings
@router.get("/{land_id}/document-role-mappings", response_model=Dict[str, Any])
async def get_project_document_role_mappings(
    land_id: UUID,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get project-specific document role mappings for a project (accessible to all authenticated users)."""
    
    # Verify land exists
    land_query = text("SELECT land_id FROM lands WHERE land_id = :land_id")
    land_result = db.execute(land_query, {"land_id": str(land_id)}).first()
    if not land_result:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found"
        )
    
    # Get all document types
    doc_types_query = text("""
        SELECT DISTINCT document_type 
        FROM documents 
        WHERE document_type IS NOT NULL
        ORDER BY document_type
    """)
    doc_types_result = db.execute(doc_types_query).fetchall()
    all_doc_types = []
    for row in doc_types_result:
        if hasattr(row, '_mapping'):
            all_doc_types.append(row._mapping['document_type'])
        elif isinstance(row, tuple):
            all_doc_types.append(row[0])
        else:
            all_doc_types.append(getattr(row, 'document_type', row[0]))
    
    # Get all reviewer roles
    roles_query = text("""
        SELECT role_key, label 
        FROM lu_roles 
        WHERE role_key IN ('re_sales_advisor', 're_analyst', 're_governance_lead')
        ORDER BY role_key
    """)
    roles_result = db.execute(roles_query).fetchall()
    all_roles = []
    for row in roles_result:
        if hasattr(row, '_mapping'):
            all_roles.append({"role_key": row._mapping['role_key'], "label": row._mapping['label']})
        elif isinstance(row, tuple):
            all_roles.append({"role_key": row[0], "label": row[1]})
        else:
            all_roles.append({"role_key": getattr(row, 'role_key', row[0]), "label": getattr(row, 'label', row[1])})
    
    # Get project-specific mappings
    mappings_query = text("""
        SELECT document_type, role_key
        FROM project_document_role_mappings
        WHERE land_id = :land_id
    """)
    try:
        mappings_result = db.execute(mappings_query, {"land_id": str(land_id)}).fetchall()
    except Exception as e:
        # Table might not exist yet, return empty mappings
        print(f"Warning: project_document_role_mappings table may not exist: {e}")
        mappings_result = []
    
    # Get default mappings
    # Try to get from document_type_roles table, or use hardcoded defaults
    default_mappings = {}
    try:
        default_mappings_query = text("""
            SELECT document_type, role_key
            FROM document_type_roles
            WHERE can_view = true
        """)
        default_mappings_result = db.execute(default_mappings_query).fetchall()
        for row in default_mappings_result:
            if hasattr(row, '_mapping'):
                doc_type = row._mapping['document_type']
                role_key = row._mapping['role_key']
            elif isinstance(row, tuple):
                doc_type = row[0]
                role_key = row[1]
            else:
                doc_type = getattr(row, 'document_type', row[0])
                role_key = getattr(row, 'role_key', row[1])
            if doc_type not in default_mappings:
                default_mappings[doc_type] = []
            default_mappings[doc_type].append(role_key)
    except Exception as e:
        # Table might not exist or have different structure, use hardcoded defaults
        print(f"Warning: document_type_roles table may not exist or have different structure: {e}")
        # Use hardcoded default mappings
        default_mappings = {
            "land-valuation": ["re_sales_advisor", "re_governance_lead"],
            "ownership-documents": ["re_governance_lead"],
            "sale-contracts": ["re_sales_advisor"],
            "topographical-surveys": ["re_sales_advisor"],
            "grid-connectivity": ["re_sales_advisor"],
            "financial-models": ["re_analyst"],
            "zoning-approvals": ["re_governance_lead"],
            "environmental-impact": ["re_governance_lead"],
            "government-nocs": ["re_governance_lead"]
        }
    
    # Build response structure: document_type -> [role_keys]
    project_mappings = {}
    for row in mappings_result:
        if hasattr(row, '_mapping'):
            doc_type = row._mapping['document_type']
            role_key = row._mapping['role_key']
        elif isinstance(row, tuple):
            doc_type = row[0]
            role_key = row[1]
        else:
            doc_type = getattr(row, 'document_type', row[0])
            role_key = getattr(row, 'role_key', row[1])
        if doc_type not in project_mappings:
            project_mappings[doc_type] = []
        project_mappings[doc_type].append(role_key)
    
    # Get user's roles to filter document types
    user_roles = current_user.get("roles", [])
    is_admin = "administrator" in user_roles
    
    # Filter document types based on user's role
    # Admin can see all document types, others see only what their roles allow
    filtered_doc_types = []
    if is_admin:
        # Admin sees all document types
        filtered_doc_types = all_doc_types if all_doc_types else [
            "land-valuation",
            "ownership-documents",
            "sale-contracts",
            "topographical-surveys",
            "grid-connectivity",
            "financial-models",
            "zoning-approvals",
            "environmental-impact",
            "government-nocs"
        ]
    else:
        # Get reviewer roles from user's roles
        reviewer_roles = [role for role in user_roles if role in ['re_sales_advisor', 're_analyst', 're_governance_lead']]
        
        # Determine which mappings to use (project-specific or default)
        effective_mappings = project_mappings if project_mappings else default_mappings
        
        # Get all document types the user's roles can access
        allowed_doc_types = set()
        for doc_type, role_keys in effective_mappings.items():
            # Check if any of the user's reviewer roles can access this document type
            if any(role in role_keys for role in reviewer_roles):
                allowed_doc_types.add(doc_type)
        
        # Filter the document types list to only include allowed ones
        # Also exclude "subtask_document" as it's not a real document type for display
        doc_types_list = all_doc_types if all_doc_types else [
            "land-valuation",
            "ownership-documents",
            "sale-contracts",
            "topographical-surveys",
            "grid-connectivity",
            "financial-models",
            "zoning-approvals",
            "environmental-impact",
            "government-nocs"
        ]
        
        filtered_doc_types = [
            dt for dt in doc_types_list 
            if dt in allowed_doc_types and dt != "subtask_document"
        ]
        
        # If user has no reviewer roles, return empty list
        if not reviewer_roles:
            filtered_doc_types = []
    
    print(f"📋 Filtered document types for user {current_user.get('user_id')} (roles: {user_roles}): {filtered_doc_types}")
    
    return {
        "land_id": str(land_id),
        "document_types": filtered_doc_types,
        "roles": all_roles,
        "project_mappings": project_mappings,
        "default_mappings": default_mappings
    }


@router.post("/{land_id}/document-role-mappings", response_model=MessageResponse)
async def save_project_document_role_mappings(
    land_id: UUID,
    request: DocumentRoleMappingsRequest,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Save project-specific document role mappings (accessible to all authenticated users).
    
    Request body format: {
        "mappings": {
            "document_type": ["role_key1", "role_key2"],
            ...
        }
    }
    """
    mappings = request.mappings
    
    # Verify land exists
    land_query = text("SELECT land_id FROM lands WHERE land_id = :land_id")
    land_result = db.execute(land_query, {"land_id": str(land_id)}).first()
    if not land_result:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found"
        )
    
    try:
        # Check if table exists first
        table_check_query = text("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables 
                WHERE table_schema = 'public' 
                AND table_name = 'project_document_role_mappings'
            )
        """)
        table_exists = db.execute(table_check_query).scalar()
        
        if not table_exists:
            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail="The project_document_role_mappings table does not exist. Please run the database migration script (renew-sql.sql) to create the required table."
            )
        
        # Delete existing mappings for this project
        delete_query = text("""
            DELETE FROM project_document_role_mappings
            WHERE land_id = :land_id
        """)
        db.execute(delete_query, {"land_id": str(land_id)})
        
        # Insert new mappings
        insert_query = text("""
            INSERT INTO project_document_role_mappings 
            (land_id, document_type, role_key, created_by)
            VALUES (:land_id, :document_type, :role_key, :created_by)
        """)
        
        for document_type, role_keys in mappings.items():
            for role_key in role_keys:
                # Verify role exists
                role_check = text("SELECT role_key FROM lu_roles WHERE role_key = :role_key")
                role_result = db.execute(role_check, {"role_key": role_key}).first()
                if not role_result:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail=f"Invalid role: {role_key}"
                    )
                
                db.execute(insert_query, {
                    "land_id": str(land_id),
                    "document_type": document_type,
                    "role_key": role_key,
                    "created_by": current_user.get("user_id")
                })
        
        db.commit()
        
        # Log the successful save for debugging
        print(f"✅ Project document role mappings saved for land_id: {land_id}")
        print(f"   Mappings saved: {len(mappings)} document types configured")
        
        return MessageResponse(
            message="Document role mappings saved successfully. Changes will be applied immediately to all document queries.",
            success=True
        )
        
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        error_msg = str(e)
        if "does not exist" in error_msg or "UndefinedTable" in error_msg:
            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail="The project_document_role_mappings table does not exist. Please run the database migration script (renew-sql.sql) to create the required table."
            )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error saving mappings: {error_msg}"
        )


@router.post("/{land_id}/site-image")
async def upload_site_image(
    land_id: UUID,
    image: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Upload site image for a project (reviewer or admin only)."""
    # Check if land exists and user has permission
    land_check = text("""
        SELECT land_id, status, landowner_id
        FROM lands 
        WHERE land_id = :land_id
    """)
    
    land_result = db.execute(land_check, {"land_id": str(land_id)}).fetchone()
    
    if not land_result:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Land not found"
        )
    
    # Check permissions: admin, reviewer, or landowner
    user_roles = current_user.get("roles", [])
    is_admin = "administrator" in user_roles
    is_reviewer = any(role in ["re_sales_advisor", "re_analyst", "re_governance_lead"] for role in user_roles)
    is_owner = str(land_result.landowner_id) == str(current_user["user_id"])
    
    if not (is_admin or is_reviewer or is_owner):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions to upload site image"
        )
    
    # Validate file type
    if not image.content_type or not image.content_type.startswith('image/'):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File must be an image"
        )
    
    # Read image data
    try:
        image_data = await image.read()
        
        # Validate file size (max 10MB)
        if len(image_data) > 10 * 1024 * 1024:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Image size must be less than 10MB"
            )
        
        # Update or insert site image
        update_query = text("""
            UPDATE lands 
            SET site_image = :image_data,
                updated_at = CURRENT_TIMESTAMP
            WHERE land_id = :land_id
            RETURNING land_id
        """)
        
        result = db.execute(update_query, {
            "land_id": str(land_id),
            "image_data": image_data
        }).fetchone()
        
        db.commit()
        
        if not result:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to upload site image"
            )
        
        return {
            "message": "Site image uploaded successfully",
            "land_id": str(land_id),
            "image_url": f"/api/lands/{land_id}/site-image"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to upload site image: {str(e)}"
        )


@router.get("/{land_id}/site-image")
async def get_site_image(
    land_id: UUID,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security),
    db: Session = Depends(get_db)
):
    """Get site image for a project (public for published projects)."""
    # Check if land exists
    query = text("""
        SELECT site_image, status, landowner_id
        FROM lands 
        WHERE land_id = :land_id
    """)
    
    result = db.execute(query, {"land_id": str(land_id)}).fetchone()
    
    if not result:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Land not found"
        )
    
    # Check permissions: public access for published projects, otherwise require auth
    is_published = result.status == "published"
    current_user = None
    
    if not is_published:
        # For non-published projects, require authentication
        if not credentials:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Authentication required"
            )
        
        try:
            from auth import verify_token, get_user_by_id
            token_data = verify_token(credentials.credentials)
            if token_data:
                current_user = get_user_by_id(token_data.user_id, db)
        except Exception:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Invalid authentication credentials"
            )
        
        if not current_user:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Authentication required"
            )
        
        user_roles = current_user.get("roles", []) if isinstance(current_user, dict) else []
        is_admin = "administrator" in user_roles
        is_reviewer = any(role in ["re_sales_advisor", "re_analyst", "re_governance_lead"] for role in user_roles)
        user_id = current_user.get("user_id") if isinstance(current_user, dict) else None
        is_owner = user_id and str(result.landowner_id) == str(user_id)
        
        if not (is_admin or is_reviewer or is_owner):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Not enough permissions to view site image"
            )
    
    if not result.site_image:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Site image not found"
        )
    
    # Return image as response
    return Response(
        content=result.site_image,
        media_type="image/jpeg",
        headers={
            "Content-Disposition": f"inline; filename=site-image-{land_id}.jpg"
        }
    )


@router.get("/{land_id}/site-image-url")
async def get_site_image_url(
    land_id: UUID,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Get site image URL for a project (returns URL if exists, null otherwise)."""
    # Check if land exists and user has permission
    query = text("""
        SELECT site_image, status, landowner_id
        FROM lands 
        WHERE land_id = :land_id
    """)
    
    result = db.execute(query, {"land_id": str(land_id)}).fetchone()
    
    if not result:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Land not found"
        )
    
    # Check permissions
    user_roles = current_user.get("roles", [])
    is_admin = "administrator" in user_roles
    is_reviewer = any(role in ["re_sales_advisor", "re_analyst", "re_governance_lead"] for role in user_roles)
    is_owner = str(result.landowner_id) == str(current_user["user_id"])
    is_published = result.status == "published"
    
    if not (is_admin or is_reviewer or is_owner or is_published):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions to view site image"
        )
    
    if result.site_image:
        return {
            "image_url": f"/api/lands/{land_id}/site-image",
            "has_image": True
        }
    else:
        return {
            "image_url": None,
            "has_image": False
        }


@router.delete("/{land_id}/site-image")
async def delete_site_image(
    land_id: UUID,
    current_user: dict = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Delete site image for a project (reviewer or admin only)."""
    # Check if land exists and user has permission
    land_check = text("""
        SELECT land_id, status, landowner_id
        FROM lands 
        WHERE land_id = :land_id
    """)
    
    land_result = db.execute(land_check, {"land_id": str(land_id)}).fetchone()
    
    if not land_result:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Land not found"
        )
    
    # Check permissions: admin or reviewer
    user_roles = current_user.get("roles", [])
    is_admin = "administrator" in user_roles
    is_reviewer = any(role in ["re_sales_advisor", "re_analyst", "re_governance_lead"] for role in user_roles)
    
    if not (is_admin or is_reviewer):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Not enough permissions to delete site image"
        )
    
    # Delete site image
    update_query = text("""
        UPDATE lands 
        SET site_image = NULL,
            updated_at = CURRENT_TIMESTAMP
        WHERE land_id = :land_id
        RETURNING land_id
    """)
    
    result = db.execute(update_query, {
        "land_id": str(land_id)
    }).fetchone()
    
    db.commit()
    
    if not result:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to delete site image"
        )
    
    return {
        "message": "Site image deleted successfully",
        "land_id": str(land_id)
    }


@router.get("/admin/reports/all-lands")
async def get_all_lands_for_report(
    current_user: dict = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Get all land details for report generation (admin only)."""
    
    query = text("""
        SELECT 
            l.land_id,
            l.title,
            l.location_text,
            l.land_type,
            l.energy_key,
            l.capacity_mw,
            l.price_per_mwh,
            l.area_acres,
            l.status,
            l.timeline_text,
            l.contract_term_years,
            l.developer_name,
            l.admin_notes,
            l.project_priority,
            l.project_due_date,
            l.created_at,
            l.updated_at,
            l.published_at,
            u.email as landowner_email,
            u.first_name,
            u.last_name,
            u.phone,
            COALESCE(COUNT(DISTINCT ii.interest_id), 0) as investor_interest_count
        FROM lands l
        LEFT JOIN "user" u ON l.landowner_id = u.user_id
        LEFT JOIN investor_interests ii ON l.land_id = ii.land_id
            AND ii.status IN ('pending', 'approved')
            AND (ii.withdrawal_requested = FALSE OR ii.withdrawal_status IS NULL OR ii.withdrawal_status != 'approved')
        WHERE l.status != 'draft'
        GROUP BY l.land_id, l.title, l.location_text, l.land_type, l.energy_key,
                 l.capacity_mw, l.price_per_mwh, l.area_acres, l.status,
                 l.timeline_text, l.contract_term_years, l.developer_name,
                 l.admin_notes, l.project_priority, l.project_due_date,
                 l.created_at, l.updated_at, l.published_at,
                 u.email, u.first_name, u.last_name, u.phone
        ORDER BY l.created_at DESC
    """)
    
    results = db.execute(query).fetchall()
    
    lands = []
    for row in results:
        landowner_name = f"{row.first_name or ''} {row.last_name or ''}".strip() or row.landowner_email or "Unknown"
        
        land = {
            "land_id": str(row.land_id),
            "title": row.title or "Untitled Project",
            "location": row.location_text or "Not specified",
            "land_type": row.land_type or "Not specified",
            "energy_type": row.energy_key or "Not specified",
            "capacity_mw": float(row.capacity_mw) if row.capacity_mw else 0,
            "price_per_mwh": float(row.price_per_mwh) if row.price_per_mwh else 0,
            "area_acres": float(row.area_acres) if row.area_acres else 0,
            "status": row.status or "unknown",
            "timeline": row.timeline_text or "Not specified",
            "contract_term_years": row.contract_term_years or 0,
            "developer_name": row.developer_name or "Not specified",
            "admin_notes": row.admin_notes or "",
            "project_priority": row.project_priority or "Not specified",
            "project_due_date": row.project_due_date.isoformat() if row.project_due_date else None,
            "created_at": row.created_at.isoformat() if row.created_at else None,
            "updated_at": row.updated_at.isoformat() if row.updated_at else None,
            "published_at": row.published_at.isoformat() if row.published_at else None,
            "landowner_name": landowner_name,
            "landowner_email": row.landowner_email or "Not specified",
            "landowner_phone": row.phone or "Not specified",
            "investor_interest_count": int(row.investor_interest_count) if row.investor_interest_count else 0
        }
        lands.append(land)
    
    return lands


@router.get("/admin/reports/excel")
async def generate_excel_report(
    current_user: dict = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Generate Excel report of all lands (admin only)."""
    
    # Get all lands data
    lands_query = text("""
        SELECT 
            l.land_id,
            l.title,
            l.location_text,
            l.land_type,
            l.energy_key,
            l.capacity_mw,
            l.price_per_mwh,
            l.area_acres,
            l.status,
            l.timeline_text,
            l.contract_term_years,
            l.developer_name,
            l.admin_notes,
            l.project_priority,
            l.project_due_date,
            l.created_at,
            l.updated_at,
            l.published_at,
            u.email as landowner_email,
            u.first_name,
            u.last_name,
            u.phone,
            COALESCE(COUNT(DISTINCT ii.interest_id), 0) as investor_interest_count
        FROM lands l
        LEFT JOIN "user" u ON l.landowner_id = u.user_id
        LEFT JOIN investor_interests ii ON l.land_id = ii.land_id
            AND ii.status IN ('pending', 'approved')
            AND (ii.withdrawal_requested = FALSE OR ii.withdrawal_status IS NULL OR ii.withdrawal_status != 'approved')
        WHERE l.status != 'draft'
        GROUP BY l.land_id, l.title, l.location_text, l.land_type, l.energy_key,
                 l.capacity_mw, l.price_per_mwh, l.area_acres, l.status,
                 l.timeline_text, l.contract_term_years, l.developer_name,
                 l.admin_notes, l.project_priority, l.project_due_date,
                 l.created_at, l.updated_at, l.published_at,
                 u.email, u.first_name, u.last_name, u.phone
        ORDER BY l.created_at DESC
    """)
    
    results = db.execute(lands_query).fetchall()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "All Lands Report"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Column headers
    headers = [
        "Land ID", "Title", "Location", "Land Type", "Energy Type",
        "Capacity (MW)", "Price per MWh", "Area (Acres)", "Status",
        "Timeline", "Contract Term (Years)", "Developer Name",
        "Project Priority", "Due Date", "Created At", "Updated At",
        "Published At", "Landowner Name", "Landowner Email", "Landowner Phone",
        "Investor Interest Count"
    ]
    
    ws.append(headers)
    
    # Apply header styling
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    for row in results:
        landowner_name = f"{row.first_name or ''} {row.last_name or ''}".strip() or row.landowner_email or "Unknown"
        
        data_row = [
            str(row.land_id),
            row.title or "Untitled Project",
            row.location_text or "Not specified",
            row.land_type or "Not specified",
            row.energy_key or "Not specified",
            float(row.capacity_mw) if row.capacity_mw else 0,
            float(row.price_per_mwh) if row.price_per_mwh else 0,
            float(row.area_acres) if row.area_acres else 0,
            row.status or "unknown",
            row.timeline_text or "Not specified",
            row.contract_term_years or 0,
            row.developer_name or "Not specified",
            row.project_priority or "Not specified",
            row.project_due_date.strftime("%Y-%m-%d") if row.project_due_date else "",
            row.created_at.strftime("%Y-%m-%d %H:%M:%S") if row.created_at else "",
            row.updated_at.strftime("%Y-%m-%d %H:%M:%S") if row.updated_at else "",
            row.published_at.strftime("%Y-%m-%d %H:%M:%S") if row.published_at else "",
            landowner_name,
            row.landowner_email or "Not specified",
            row.phone or "Not specified",
            int(row.investor_interest_count) if row.investor_interest_count else 0
        ]
        ws.append(data_row)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"lands_report_{timestamp}.xlsx"
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/admin/reports/pdf")
async def generate_pdf_report(
    current_user: dict = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Generate PDF report of all lands (admin only)."""
    
    # Get all lands data
    lands_query = text("""
        SELECT 
            l.land_id,
            l.title,
            l.location_text,
            l.land_type,
            l.energy_key,
            l.capacity_mw,
            l.price_per_mwh,
            l.area_acres,
            l.status,
            l.timeline_text,
            l.contract_term_years,
            l.developer_name,
            l.admin_notes,
            l.project_priority,
            l.project_due_date,
            l.created_at,
            l.updated_at,
            l.published_at,
            u.email as landowner_email,
            u.first_name,
            u.last_name,
            u.phone,
            COALESCE(COUNT(DISTINCT ii.interest_id), 0) as investor_interest_count
        FROM lands l
        LEFT JOIN "user" u ON l.landowner_id = u.user_id
        LEFT JOIN investor_interests ii ON l.land_id = ii.land_id
            AND ii.status IN ('pending', 'approved')
            AND (ii.withdrawal_requested = FALSE OR ii.withdrawal_status IS NULL OR ii.withdrawal_status != 'approved')
        WHERE l.status != 'draft'
        GROUP BY l.land_id, l.title, l.location_text, l.land_type, l.energy_key,
                 l.capacity_mw, l.price_per_mwh, l.area_acres, l.status,
                 l.timeline_text, l.contract_term_years, l.developer_name,
                 l.admin_notes, l.project_priority, l.project_due_date,
                 l.created_at, l.updated_at, l.published_at,
                 u.email, u.first_name, u.last_name, u.phone
        ORDER BY l.created_at DESC
    """)
    
    results = db.execute(lands_query).fetchall()
    
    # Create PDF buffer
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    # Container for PDF elements
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#366092'),
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    # Title
    title = Paragraph("All Lands Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Summary paragraph
    summary_text = f"<b>Generated on:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>"
    summary_text += f"<b>Total Projects:</b> {len(results)}"
    summary = Paragraph(summary_text, styles['Normal'])
    elements.append(summary)
    elements.append(Spacer(1, 0.3*inch))
    
    # Prepare table data
    table_data = []
    
    # Table headers
    headers = [
        "Title", "Location", "Energy Type", "Capacity (MW)",
        "Status", "Landowner", "Created"
    ]
    table_data.append(headers)
    
    # Table rows (limit to prevent PDF from being too large)
    max_rows = 100  # Limit rows per page
    for idx, row in enumerate(results[:max_rows]):
        landowner_name = f"{row.first_name or ''} {row.last_name or ''}".strip() or row.landowner_email or "Unknown"
        
        data_row = [
            row.title or "Untitled Project",
            row.location_text or "Not specified",
            row.energy_key or "Not specified",
            f"{float(row.capacity_mw):.2f}" if row.capacity_mw else "0",
            row.status or "unknown",
            landowner_name[:30],  # Truncate long names
            row.created_at.strftime("%Y-%m-%d") if row.created_at else ""
        ]
        table_data.append(data_row)
    
    # Create table
    table = Table(table_data, colWidths=[1.5*inch, 1.5*inch, 1*inch, 0.8*inch, 1*inch, 1.2*inch, 0.8*inch])
    
    # Table style
    table.setStyle(TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        # Data rows
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    
    # Add note if rows were truncated
    if len(results) > max_rows:
        note = Paragraph(
            f"<i>Note: Showing first {max_rows} of {len(results)} projects. Use Excel export for complete data.</i>",
            styles['Normal']
        )
        elements.append(Spacer(1, 0.2*inch))
        elements.append(note)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"lands_report_{timestamp}.pdf"
    
    return StreamingResponse(
        io.BytesIO(buffer.read()),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )